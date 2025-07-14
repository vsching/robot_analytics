"""
Export Manager

Comprehensive export and reporting functionality for trading strategy analysis data.
Supports CSV, Excel, and PDF formats with customizable templates and scheduling.
"""

import os
import csv
import io
import logging
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Union, Tuple
from dataclasses import dataclass, asdict
from decimal import Decimal
import pandas as pd
import numpy as np
from pathlib import Path

# Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF support
try:
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_pdf import PdfPages
    import seaborn as sns
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from ..models import Trade, Strategy
from ..analytics import PerformanceMetrics, ConfluenceMetrics, SignalOverlap
from ..services.strategy_manager import StrategyManager
from ..analytics import AnalyticsEngine


logger = logging.getLogger(__name__)


@dataclass
class ExportConfig:
    """Configuration for export operations."""
    include_trades: bool = True
    include_metrics: bool = True
    include_monthly_breakdown: bool = True
    include_confluence: bool = False
    date_format: str = "%Y-%m-%d"
    decimal_places: int = 2
    delimiter: str = ","
    include_headers: bool = True
    custom_columns: Optional[List[str]] = None


@dataclass
class ReportTemplate:
    """Template configuration for PDF reports."""
    title: str = "Trading Strategy Analysis Report"
    subtitle: str = ""
    logo_path: Optional[str] = None
    author: str = "Trading Strategy Analyzer"
    include_summary: bool = True
    include_charts: bool = True
    include_detailed_metrics: bool = True
    include_confluence_analysis: bool = False
    color_scheme: str = "default"  # default, professional, dark


class ExportManager:
    """Manages all export and reporting functionality."""
    
    def __init__(self, 
                 strategy_manager: StrategyManager,
                 analytics_engine: AnalyticsEngine,
                 export_dir: str = "exports"):
        """
        Initialize export manager.
        
        Args:
            strategy_manager: Strategy manager instance
            analytics_engine: Analytics engine instance
            export_dir: Directory for exported files
        """
        self.strategy_manager = strategy_manager
        self.analytics_engine = analytics_engine
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)
        
        # Create subdirectories
        (self.export_dir / "csv").mkdir(exist_ok=True)
        (self.export_dir / "excel").mkdir(exist_ok=True)
        (self.export_dir / "pdf").mkdir(exist_ok=True)
    
    def export_trades_to_csv(self, 
                           strategy_id: Optional[int] = None,
                           config: Optional[ExportConfig] = None) -> str:
        """
        Export trades to CSV format.
        
        Args:
            strategy_id: Strategy ID to export (None for all)
            config: Export configuration
            
        Returns:
            Path to exported file
        """
        config = config or ExportConfig()
        
        # Get trades data
        if strategy_id:
            trades = self.analytics_engine.get_trades_for_strategy(strategy_id)
            strategy = self.strategy_manager.get_strategy(strategy_id)
            filename = f"trades_{strategy.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            strategies = self.strategy_manager.get_all_strategies()
            trades = []
            for strategy in strategies:
                strategy_trades = self.analytics_engine.get_trades_for_strategy(strategy.id)
                trades.extend(strategy_trades)
            filename = f"all_trades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        if not trades:
            raise ValueError("No trades found to export")
        
        # Prepare data
        export_data = []
        for trade in trades:
            row = {
                'strategy_id': trade.strategy_id,
                'trade_date': trade.trade_date.strftime(config.date_format),
                'symbol': trade.symbol or '',
                'side': trade.side or '',
                'entry_price': round(float(trade.entry_price or 0), config.decimal_places),
                'exit_price': round(float(trade.exit_price or 0), config.decimal_places),
                'quantity': round(float(trade.quantity or 0), config.decimal_places),
                'pnl': round(float(trade.pnl or 0), config.decimal_places),
                'commission': round(float(trade.commission or 0), config.decimal_places)
            }
            
            # Add strategy name
            strategy = self.strategy_manager.get_strategy(trade.strategy_id)
            row['strategy_name'] = strategy.name if strategy else f"Strategy {trade.strategy_id}"
            
            export_data.append(row)
        
        # Apply custom column filtering
        if config.custom_columns:
            filtered_data = []
            for row in export_data:
                filtered_row = {col: row.get(col, '') for col in config.custom_columns}
                filtered_data.append(filtered_row)
            export_data = filtered_data
        
        # Write CSV
        filepath = self.export_dir / "csv" / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            if export_data:
                fieldnames = list(export_data[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=config.delimiter)
                
                if config.include_headers:
                    writer.writeheader()
                
                writer.writerows(export_data)
        
        logger.info(f"Exported {len(trades)} trades to {filepath}")
        return str(filepath)
    
    def export_metrics_to_csv(self, 
                            strategy_id: Optional[int] = None,
                            config: Optional[ExportConfig] = None) -> str:
        """
        Export performance metrics to CSV format.
        
        Args:
            strategy_id: Strategy ID to export (None for all)
            config: Export configuration
            
        Returns:
            Path to exported file
        """
        config = config or ExportConfig()
        
        # Get strategies
        if strategy_id:
            strategies = [self.strategy_manager.get_strategy(strategy_id)]
            filename = f"metrics_{strategies[0].name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            strategies = self.strategy_manager.get_all_strategies()
            filename = f"all_metrics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        strategies = [s for s in strategies if s is not None]
        
        if not strategies:
            raise ValueError("No strategies found to export")
        
        # Prepare metrics data
        export_data = []
        for strategy in strategies:
            metrics = self.analytics_engine.calculate_metrics_for_strategy(strategy.id)
            if not metrics:
                continue
            
            row = {
                'strategy_id': strategy.id,
                'strategy_name': strategy.name,
                'total_pnl': round(float(metrics.pnl_summary.total_pnl), config.decimal_places),
                'total_trades': metrics.trade_statistics.total_trades,
                'winning_trades': metrics.trade_statistics.winning_trades,
                'losing_trades': metrics.trade_statistics.losing_trades,
                'win_rate': round(metrics.trade_statistics.win_rate, config.decimal_places),
                'average_win': round(float(metrics.trade_statistics.average_win), config.decimal_places),
                'average_loss': round(float(metrics.trade_statistics.average_loss), config.decimal_places),
                'profit_factor': round(metrics.trade_statistics.profit_factor, config.decimal_places),
                'max_consecutive_wins': metrics.trade_statistics.max_consecutive_wins,
                'max_consecutive_losses': metrics.trade_statistics.max_consecutive_losses,
                'largest_win': round(float(metrics.pnl_summary.max_pnl), config.decimal_places),
                'largest_loss': round(float(metrics.pnl_summary.min_pnl), config.decimal_places),
                'average_pnl': round(float(metrics.pnl_summary.average_pnl), config.decimal_places),
                'median_pnl': round(float(metrics.pnl_summary.median_pnl), config.decimal_places),
                'std_dev': round(float(metrics.pnl_summary.std_dev), config.decimal_places)
            }
            
            # Add advanced metrics if available
            if metrics.advanced_statistics:
                row.update({
                    'sharpe_ratio': round(metrics.advanced_statistics.sharpe_ratio or 0, config.decimal_places),
                    'sortino_ratio': round(metrics.advanced_statistics.sortino_ratio or 0, config.decimal_places),
                    'calmar_ratio': round(metrics.advanced_statistics.calmar_ratio or 0, config.decimal_places),
                    'max_drawdown': round(metrics.advanced_statistics.max_drawdown, config.decimal_places),
                    'max_drawdown_duration': metrics.advanced_statistics.max_drawdown_duration or 0,
                    'recovery_duration': metrics.advanced_statistics.recovery_duration or 0,
                    'value_at_risk_95': round(metrics.advanced_statistics.value_at_risk_95 or 0, config.decimal_places),
                    'conditional_var_95': round(metrics.advanced_statistics.conditional_var_95 or 0, config.decimal_places)
                })
            
            export_data.append(row)
        
        # Write CSV
        filepath = self.export_dir / "csv" / filename
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            if export_data:
                fieldnames = list(export_data[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=config.delimiter)
                
                if config.include_headers:
                    writer.writeheader()
                
                writer.writerows(export_data)
        
        logger.info(f"Exported metrics for {len(strategies)} strategies to {filepath}")
        return str(filepath)
    
    def export_to_excel(self, 
                       strategy_id: Optional[int] = None,
                       config: Optional[ExportConfig] = None,
                       include_charts: bool = True) -> str:
        """
        Export comprehensive data to Excel with multiple sheets.
        
        Args:
            strategy_id: Strategy ID to export (None for all)
            config: Export configuration
            include_charts: Whether to include charts
            
        Returns:
            Path to exported file
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        config = config or ExportConfig()
        
        # Determine filename
        if strategy_id:
            strategy = self.strategy_manager.get_strategy(strategy_id)
            filename = f"report_{strategy.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"full_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = self.export_dir / "excel" / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add summary sheet
        self._add_summary_sheet(wb, strategy_id)
        
        # Add trades sheet
        if config.include_trades:
            self._add_trades_sheet(wb, strategy_id, config)
        
        # Add metrics sheet
        if config.include_metrics:
            self._add_metrics_sheet(wb, strategy_id, config)
        
        # Add monthly breakdown sheet
        if config.include_monthly_breakdown:
            self._add_monthly_sheet(wb, strategy_id, config)
        
        # Add charts sheet
        if include_charts:
            self._add_charts_sheet(wb, strategy_id)
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"Exported Excel report to {filepath}")
        return str(filepath)
    
    def _add_summary_sheet(self, wb: openpyxl.Workbook, strategy_id: Optional[int]):
        """Add summary sheet to Excel workbook."""
        ws = wb.create_sheet("Summary")
        
        # Header
        ws['A1'] = "Trading Strategy Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Strategy information
        row = 4
        if strategy_id:
            strategy = self.strategy_manager.get_strategy(strategy_id)
            ws[f'A{row}'] = "Strategy:"
            ws[f'B{row}'] = strategy.name
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = "Description:"
            ws[f'B{row}'] = strategy.description or "No description"
            ws[f'A{row}'].font = Font(bold=True)
            row += 2
        else:
            strategies = self.strategy_manager.get_all_strategies()
            ws[f'A{row}'] = "Strategies Included:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for strategy in strategies:
                ws[f'B{row}'] = strategy.name
                row += 1
            row += 1
        
        # Key metrics summary
        ws[f'A{row}'] = "Key Performance Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Get aggregated metrics
        if strategy_id:
            metrics = self.analytics_engine.calculate_metrics_for_strategy(strategy_id)
            if metrics:
                self._add_metrics_to_sheet(ws, metrics, row)
        else:
            # Aggregate all strategies
            all_trades = []
            strategies = self.strategy_manager.get_all_strategies()
            for strategy in strategies:
                trades = self.analytics_engine.get_trades_for_strategy(strategy.id)
                all_trades.extend(trades)
            
            if all_trades:
                pnl_summary = self.analytics_engine.calculate_pnl_summary(all_trades)
                trade_stats = self.analytics_engine.calculate_trade_statistics(all_trades)
                
                self._add_summary_metrics_to_sheet(ws, pnl_summary, trade_stats, row)
    
    def _add_trades_sheet(self, wb: openpyxl.Workbook, strategy_id: Optional[int], config: ExportConfig):
        """Add trades sheet to Excel workbook."""
        ws = wb.create_sheet("Trades")
        
        # Get trades
        if strategy_id:
            trades = self.analytics_engine.get_trades_for_strategy(strategy_id)
        else:
            trades = []
            strategies = self.strategy_manager.get_all_strategies()
            for strategy in strategies:
                strategy_trades = self.analytics_engine.get_trades_for_strategy(strategy.id)
                trades.extend(strategy_trades)
        
        # Prepare DataFrame
        trade_data = []
        for trade in trades:
            strategy = self.strategy_manager.get_strategy(trade.strategy_id)
            trade_data.append({
                'Strategy': strategy.name if strategy else f"Strategy {trade.strategy_id}",
                'Date': trade.trade_date,
                'Symbol': trade.symbol or '',
                'Side': trade.side or '',
                'Entry Price': float(trade.entry_price or 0),
                'Exit Price': float(trade.exit_price or 0),
                'Quantity': float(trade.quantity or 0),
                'P&L': float(trade.pnl or 0),
                'Commission': float(trade.commission or 0)
            })
        
        if trade_data:
            df = pd.DataFrame(trade_data)
            
            # Add to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def _add_metrics_sheet(self, wb: openpyxl.Workbook, strategy_id: Optional[int], config: ExportConfig):
        """Add metrics sheet to Excel workbook."""
        ws = wb.create_sheet("Metrics")
        
        # Headers
        headers = [
            'Strategy', 'Total P&L', 'Total Trades', 'Win Rate (%)', 
            'Avg Win', 'Avg Loss', 'Profit Factor', 'Max Drawdown (%)',
            'Sharpe Ratio', 'Sortino Ratio'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Get strategies
        if strategy_id:
            strategies = [self.strategy_manager.get_strategy(strategy_id)]
        else:
            strategies = self.strategy_manager.get_all_strategies()
        
        # Add data
        row = 2
        for strategy in strategies:
            if not strategy:
                continue
                
            metrics = self.analytics_engine.calculate_metrics_for_strategy(strategy.id)
            if not metrics:
                continue
            
            ws.cell(row=row, column=1, value=strategy.name)
            ws.cell(row=row, column=2, value=float(metrics.pnl_summary.total_pnl))
            ws.cell(row=row, column=3, value=metrics.trade_statistics.total_trades)
            ws.cell(row=row, column=4, value=round(metrics.trade_statistics.win_rate, 2))
            ws.cell(row=row, column=5, value=float(metrics.trade_statistics.average_win))
            ws.cell(row=row, column=6, value=float(metrics.trade_statistics.average_loss))
            ws.cell(row=row, column=7, value=round(metrics.trade_statistics.profit_factor, 2))
            
            if metrics.advanced_statistics:
                ws.cell(row=row, column=8, value=round(metrics.advanced_statistics.max_drawdown * 100, 2))
                ws.cell(row=row, column=9, value=round(metrics.advanced_statistics.sharpe_ratio or 0, 2))
                ws.cell(row=row, column=10, value=round(metrics.advanced_statistics.sortino_ratio or 0, 2))
            
            row += 1
    
    def _add_monthly_sheet(self, wb: openpyxl.Workbook, strategy_id: Optional[int], config: ExportConfig):
        """Add monthly breakdown sheet to Excel workbook."""
        ws = wb.create_sheet("Monthly Breakdown")
        
        # Get trades and calculate monthly breakdown
        if strategy_id:
            trades = self.analytics_engine.get_trades_for_strategy(strategy_id)
        else:
            trades = []
            strategies = self.strategy_manager.get_all_strategies()
            for strategy in strategies:
                strategy_trades = self.analytics_engine.get_trades_for_strategy(strategy.id)
                trades.extend(strategy_trades)
        
        if not trades:
            ws['A1'] = "No trades found for monthly breakdown"
            return
        
        # Calculate monthly aggregations
        monthly_data = {}
        for trade in trades:
            if trade.pnl:
                month_key = trade.trade_date.strftime('%Y-%m')
                if month_key not in monthly_data:
                    monthly_data[month_key] = {
                        'trades': 0,
                        'pnl': 0,
                        'wins': 0,
                        'losses': 0
                    }
                
                monthly_data[month_key]['trades'] += 1
                monthly_data[month_key]['pnl'] += float(trade.pnl)
                if trade.pnl > 0:
                    monthly_data[month_key]['wins'] += 1
                else:
                    monthly_data[month_key]['losses'] += 1
        
        # Headers
        headers = ['Month', 'Trades', 'P&L', 'Wins', 'Losses', 'Win Rate (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        row = 2
        for month, data in sorted(monthly_data.items()):
            win_rate = (data['wins'] / data['trades'] * 100) if data['trades'] > 0 else 0
            
            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=data['trades'])
            ws.cell(row=row, column=3, value=round(data['pnl'], 2))
            ws.cell(row=row, column=4, value=data['wins'])
            ws.cell(row=row, column=5, value=data['losses'])
            ws.cell(row=row, column=6, value=round(win_rate, 2))
            
            row += 1
    
    def _add_charts_sheet(self, wb: openpyxl.Workbook, strategy_id: Optional[int]):
        """Add charts sheet to Excel workbook."""
        ws = wb.create_sheet("Charts")
        ws['A1'] = "Charts and Visualizations"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A3'] = "Note: Charts are best viewed in the web dashboard for interactive features."
    
    def _add_metrics_to_sheet(self, ws, metrics: PerformanceMetrics, start_row: int):
        """Add metrics to worksheet."""
        row = start_row
        
        # Basic metrics
        ws[f'A{row}'] = "Total P&L:"
        ws[f'B{row}'] = float(metrics.pnl_summary.total_pnl)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total Trades:"
        ws[f'B{row}'] = metrics.trade_statistics.total_trades
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Win Rate:"
        ws[f'B{row}'] = f"{metrics.trade_statistics.win_rate:.1f}%"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Profit Factor:"
        ws[f'B{row}'] = round(metrics.trade_statistics.profit_factor, 2)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        if metrics.advanced_statistics:
            ws[f'A{row}'] = "Sharpe Ratio:"
            ws[f'B{row}'] = round(metrics.advanced_statistics.sharpe_ratio or 0, 2)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = "Max Drawdown:"
            ws[f'B{row}'] = f"{metrics.advanced_statistics.max_drawdown:.1%}"
            ws[f'A{row}'].font = Font(bold=True)
    
    def _add_summary_metrics_to_sheet(self, ws, pnl_summary, trade_stats, start_row: int):
        """Add summary metrics to worksheet."""
        row = start_row
        
        ws[f'A{row}'] = "Total P&L:"
        ws[f'B{row}'] = float(pnl_summary.total_pnl)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total Trades:"
        ws[f'B{row}'] = trade_stats.total_trades
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Win Rate:"
        ws[f'B{row}'] = f"{trade_stats.win_rate:.1f}%"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Profit Factor:"
        ws[f'B{row}'] = round(trade_stats.profit_factor, 2)
        ws[f'A{row}'].font = Font(bold=True)
    
    def get_available_formats(self) -> List[str]:
        """Get list of available export formats."""
        formats = ['csv']
        
        if EXCEL_AVAILABLE:
            formats.append('excel')
        
        if PDF_AVAILABLE:
            formats.append('pdf')
        
        return formats
    
    def cleanup_old_exports(self, days_old: int = 30):
        """
        Clean up old export files.
        
        Args:
            days_old: Delete files older than this many days
        """
        cutoff_date = datetime.now() - timedelta(days=days_old)
        
        for subdir in ['csv', 'excel', 'pdf']:
            export_subdir = self.export_dir / subdir
            if export_subdir.exists():
                for file_path in export_subdir.iterdir():
                    if file_path.is_file():
                        file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                        if file_time < cutoff_date:
                            file_path.unlink()
                            logger.info(f"Deleted old export file: {file_path}")
    
    def get_export_history(self) -> List[Dict[str, Any]]:
        """
        Get history of exported files.
        
        Returns:
            List of export file information
        """
        history = []
        
        for subdir in ['csv', 'excel', 'pdf']:
            export_subdir = self.export_dir / subdir
            if export_subdir.exists():
                for file_path in export_subdir.iterdir():
                    if file_path.is_file():
                        stat = file_path.stat()
                        history.append({
                            'filename': file_path.name,
                            'format': subdir,
                            'size_bytes': stat.st_size,
                            'created': datetime.fromtimestamp(stat.st_ctime),
                            'modified': datetime.fromtimestamp(stat.st_mtime),
                            'path': str(file_path)
                        })
        
        return sorted(history, key=lambda x: x['modified'], reverse=True)